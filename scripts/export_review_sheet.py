#!/usr/bin/env python3
"""Render specification, implementation and live evidence into one review sheet.

The core keeps its truth in two places, written by two different sides and never
derived from one another:

  * ``erp-spec.yaml``   — what the ERP must be able to do and to record (domain expert)
  * ``verified.json``   — what a live run actually demonstrated (the prober)

Reviewing either alone answers the wrong question. The spec alone cannot say whether
anything works; ``verified.json`` alone cannot say whether what works is what was
needed. This joins them onto one row per capability and per field requirement, so the
review question becomes "is anything missing for the business?" instead of "does this
key exist?".

Two outputs, same data:

  cores/<id>/review.yaml   diffable in a PR — a change here means something moved
  cores/<id>/review.xlsx   the sheet a domain expert reads

The YAML deliberately carries NO generation timestamp. It records the stamp of the
probe run it was built from instead, so the file changes only when the underlying
facts change. A file that churns on every run cannot be reviewed by diff, which is
the only reason to check a generated artefact in at all.

Rows are ordered along the process chain (customer → quote → order → delivery →
invoice → return → credit note, then purchasing, then master data), because that is
how the business reads its own system — an alphabetical list puts `CreditNote` three
screens away from the `Return` that produced it.

Run (openpyxl is deliberately not a repo dependency — supplied per call, like the
test runs supply pytest):

    PYTHONPATH=<agent-os>/backend \\
      uv run --project <agent-os>/backend --with openpyxl --with pyyaml \\
      python scripts/export_review_sheet.py [core_id]
"""

from __future__ import annotations

import json
import sys
import types
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parent.parent
_CORES = _ROOT / "cores"

# Cores import as `xentral_entity_cores.<id>` (never top-level — a core id like
# `xentral_api` would shadow a same-named backend module). Mirrors conftest.py.
_PKG = "xentral_entity_cores"
if _PKG not in sys.modules:
    _pkg = types.ModuleType(_PKG)
    _pkg.__package__ = _PKG
    _pkg.__path__ = [str(_CORES)]  # type: ignore[attr-defined]
    sys.modules[_PKG] = _pkg

import yaml  # noqa: E402

# openpyxl is imported inside `_write_xlsx`, not here. It is deliberately not a repo
# dependency (supplied per call), and the drift test rebuilds the YAML half through
# `render_yaml` — which must stay importable on a plain test runner.

#: The verdict that means the capability itself was demonstrated — an effect read
#: back, not merely a route that answered. Mirrors ``verdicts.PROVEN``.
PROVEN = "pass"

# How each verdict reads on the sheet. The distinction between "proven" and
# "reachable" is the entire point of the column: `reachable` means the route exists
# and refused our probe, which is no capability claim in either direction.
#: Machine token → how it reads on the sheet. The rows carry the tokens (the YAML is
#: diffed by tools); the German wording exists only here, where a person reads it.
REQUIREMENT_LABEL = {True: "muss funktionieren", False: "Lücke akzeptiert"}
IMPLEMENTATION_LABEL = {
    "built": "gebaut",
    "blockedUpstream": "nicht möglich (Upstream)",
    "missing": "fehlt",
}
FIELD_STATUS_LABEL = {
    "open": "offen",
    "recheck": "offen — prüfen, evtl. erledigt",
    "fieldMissing": "Feld fehlt im Modell",
}

VERDICT_LABEL = {
    PROVEN: "bewiesen",
    "executed": "ausgeführt, nicht nachgelesen",
    "accepted": "akzeptiert, nicht nachgelesen",
    "reachable": "nur erreichbar — kein Nachweis",
    "unobserved": "keine Daten im Sample",
    "fail": "FEHLER",
    None: "ungetestet",
}

# op → the schema flag that says whether the field supports it at all. `read` has no
# flag: everything declared is readable, including readOnly fields.
OP_FLAG = {
    "create": "creatable",
    "update": "updatable",
    "filter": "filterable",
    "sort": "sortable",
    "search": "searchable",
}

# Cell colours, as hex — the PatternFill objects are built inside `_write_xlsx`,
# which is the only place openpyxl exists.
GREEN, AMBER, RED, GREY, BLUE, HEAD = (
    "C6E0B4",
    "FFE699",
    "F4B183",
    "EDEDED",
    "BDD7EE",
    "44546A",
)


def _load_core(core_id: str):
    module = __import__(f"xentral_entity_cores.{core_id}.manifest", fromlist=["CORE"])
    return module.CORE


def _walk(props: dict[str, Any] | None, prefix: str = ""):
    for name, spec in (props or {}).items():
        if not isinstance(spec, dict):
            continue
        path = f"{prefix}{name}"
        yield path, spec
        node = spec.get("node")
        nested = node.get("properties") if isinstance(node, dict) else spec.get("properties")
        if isinstance(nested, dict):
            yield from _walk(nested, f"{path}.")


def _model(core_id: str) -> dict[str, dict[str, Any]]:
    """Every adapter's metadata, indexed the way the spec addresses it.

    Actions and step commands share one namespace here: a capability key is unique
    across both (the playbook test asserts it), and a reviewer does not care which
    of the two schema slots a capability happens to sit in.
    """
    out: dict[str, dict[str, Any]] = {}
    for adapter in _load_core(core_id).adapters:
        meta = adapter.metadata(None)
        capabilities: dict[str, dict[str, Any]] = {}
        for action in meta.get("actions") or []:
            if action.get("key"):
                capabilities[action["key"]] = action
        for group in meta.get("processSteps") or []:
            for command in group.get("commands") or []:
                if command.get("key"):
                    capabilities[command["key"]] = command
        out[meta["key"]] = {
            "label": meta.get("label") or meta["key"],
            "operations": list(meta.get("operations") or []),
            "capabilities": capabilities,
            "fields": dict(_walk((meta.get("rootNode") or {}).get("properties"))),
        }
    return out


def _order(core_id: str, keys) -> list[str]:
    """The core's own reading sequence. Shared with the spec and its ordering rule
    (`cores/<id>/order.py`) so the sheet and the file a reviewer edits agree."""
    module = __import__(f"xentral_entity_cores.{core_id}.order", fromlist=["chain_order"])
    return module.chain_order(keys)


def _capability_rows(core_id: str, spec: dict, model: dict, verified: dict) -> list[dict[str, Any]]:
    """One row per capability: what the spec asks, what the core built, what a run proved."""
    rows: list[dict[str, Any]] = []
    for entity in _order(core_id, set(spec) | set(model)):
        block = spec.get(entity) or {}
        meta = model.get(entity) or {"capabilities": {}}
        marks = verified.get(entity) or {}
        can, cannot = block.get("can") or {}, block.get("cannot") or {}
        for key in sorted(set(can) | set(cannot)):
            capability = meta["capabilities"].get(key) or {}
            required = key in can
            wish = capability.get("wish")
            verdict = marks.get(key) if required else None
            rows.append(
                {
                    "entity": entity,
                    "capability": key,
                    "label": capability.get("label") or key,
                    # What the spec asks of the core, and what the core answers.
                    # Machine tokens: the YAML is diffed by tools, the German
                    # wording lives in the sheet writer where a human reads it.
                    "required": required,
                    "implementation": (
                        "missing" if not capability else "blockedUpstream" if wish else "built"
                    ),
                    # A gap carries no evidence to report — nothing was expected to run.
                    "verdict": verdict,
                    "proven": bool(required and verdict == PROVEN),
                    "note": wish or "",
                    "reviewed": block.get("reviewed") or "",
                }
            )
    return rows


def _field_rows(core_id: str, spec: dict, model: dict, verified: dict) -> list[dict[str, Any]]:
    """One row per field requirement — the other axis of the same specification.

    A recorded gap can go stale, and a stale one is worse than none: it goes on
    claiming "not possible" over a capability that has since been built, and it wins
    over a live pass wherever it is shown. So each row also reports whether the
    schema now declares the flag and whether a run has proven that facet — that is
    what makes an obsolete entry visible instead of permanent.
    """
    rows: list[dict[str, Any]] = []
    # The gaps sit on the field they concern now; flattened here to the row shape.
    entities = {}
    for key, block in spec.items():
        rows = [
            {"field": path, **gap}
            for path, field in (block.get("fields") or {}).items()
            for gap in (field.get("gaps") or [])
        ]
        if rows:
            entities[key] = rows
    for entity in _order(core_id, entities):
        meta = model.get(entity)
        marks = (verified.get(entity) or {}).get("fields") or {}
        for entry in entities[entity] or []:
            path = entry.get("field") or ""
            ops = list(entry.get("ops") or [])
            spec_field = (meta or {}).get("fields", {}).get(path)
            declared = [op for op in ops if op in OP_FLAG and (spec_field or {}).get(OP_FLAG[op])]
            proven = [op for op in ops if (marks.get(path) or {}).get(op) == PROVEN]
            rows.append(
                {
                    "entity": entity,
                    "field": path,
                    "ops": ops,
                    "reason": entry.get("reason") or "",
                    "fieldExists": spec_field is not None,
                    # Declared writable/filterable despite the recorded gap, or
                    # outright proven: either way the entry needs a second look.
                    "nowDeclared": declared,
                    "nowProven": proven,
                    "status": (
                        "fieldMissing"
                        if spec_field is None
                        else "recheck"
                        if (declared or proven)
                        else "open"
                    ),
                }
            )
    return rows


def _summary(
    core_id: str, cap_rows: list[dict], field_rows: list[dict], spec: dict
) -> list[dict[str, Any]]:
    reviewed = {key: block.get("reviewed") for key, block in spec.items()}
    per: dict[str, dict[str, Any]] = {}
    for row in cap_rows:
        s = per.setdefault(
            row["entity"],
            {
                "entity": row["entity"],
                "required": 0,
                "proven": 0,
                "unproven": 0,
                "acceptedGaps": 0,
                "notBuilt": 0,
                "fieldRequirements": 0,
                "reviewed": reviewed.get(row["entity"]) or "",
            },
        )
        if row["required"]:
            s["required"] += 1
            s["proven" if row["proven"] else "unproven"] += 1
        else:
            s["acceptedGaps"] += 1
        if row["implementation"] == "missing":
            s["notBuilt"] += 1
    for row in field_rows:
        s = per.setdefault(
            row["entity"],
            {
                "entity": row["entity"],
                "required": 0,
                "proven": 0,
                "unproven": 0,
                "acceptedGaps": 0,
                "notBuilt": 0,
                "fieldRequirements": 0,
                "reviewed": reviewed.get(row["entity"]) or "",
            },
        )
        s["fieldRequirements"] += 1
    return [per[k] for k in _order(core_id, per)]


def render_yaml(core_id: str, source: dict, summary, cap_rows, field_rows) -> str:
    """The diffable half, as text. No own timestamp — see the module docstring.

    Split out from writing so the drift test can rebuild it and compare against the
    checked-in file without needing openpyxl, which is not a repo dependency.
    """
    doc = {
        "core": core_id,
        "source": source,
        "totals": {
            "required": sum(s["required"] for s in summary),
            "proven": sum(s["proven"] for s in summary),
            "unproven": sum(s["unproven"] for s in summary),
            "acceptedGaps": sum(s["acceptedGaps"] for s in summary),
            "fieldRequirements": sum(s["fieldRequirements"] for s in summary),
            "reviewedEntities": sum(1 for s in summary if s["reviewed"]),
        },
        "entities": {
            s["entity"]: {
                "reviewed": s["reviewed"] or None,
                "capabilities": [
                    {
                        "key": r["capability"],
                        "requirement": "required" if r["required"] else "acceptedGap",
                        "implementation": r["implementation"],
                        "evidence": r["verdict"],
                        "proven": r["proven"],
                        **({"reason": r["note"]} if r["note"] else {}),
                    }
                    for r in cap_rows
                    if r["entity"] == s["entity"]
                ],
                "fieldRequirements": [
                    {
                        "field": r["field"],
                        "ops": r["ops"],
                        "status": r["status"],
                        "reason": r["reason"],
                        **({"nowDeclared": r["nowDeclared"]} if r["nowDeclared"] else {}),
                        **({"nowProven": r["nowProven"]} if r["nowProven"] else {}),
                    }
                    for r in field_rows
                    if r["entity"] == s["entity"]
                ],
            }
            for s in summary
        },
    }
    header = (
        "# Review sheet — specification vs. implementation vs. live evidence.\n"
        "# GENERATED by scripts/export_review_sheet.py. Do not edit: change the\n"
        "# specification (erp-spec.yaml) or the code.\n"
        "# Carries no generation timestamp on purpose — `source` records the probe run\n"
        "# it was built from, so a diff here means a fact moved, not that it was re-run.\n"
    )
    return header + yaml.safe_dump(doc, allow_unicode=True, sort_keys=False)


def build(core_id: str) -> tuple[dict, list[dict], list[dict], list[dict]]:
    """Read the three sources and assemble the review rows. No output side."""
    core_dir = _CORES / core_id
    spec_path = core_dir / "erp-spec.yaml"
    # category → entity → block; every consumer here addresses an entity, so the
    # grouping is flattened once (same seam as the core's own loader).
    grouped = yaml.safe_load(spec_path.read_text("utf-8")) or {} if spec_path.is_file() else {}
    spec = {k: b for entities in grouped.values() for k, b in entities.items()}

    verified_path = core_dir / "verified.json"
    verified_doc = json.loads(verified_path.read_text("utf-8")) if verified_path.is_file() else {}
    # Actions and step commands are separate slots in the manifest but one namespace
    # to a reviewer; fields stay nested so the field axis can address them by path.
    verified = {
        key: {
            **(entry.get("actions") or {}),
            **(entry.get("processSteps") or {}),
            "fields": entry.get("fields") or {},
        }
        for key, entry in (verified_doc.get("entities") or {}).items()
    }
    source = {
        "probedAt": verified_doc.get("generatedAt"),
        "instance": verified_doc.get("instance"),
    }

    model = _model(core_id)
    cap_rows = _capability_rows(core_id, spec, model, verified)
    field_rows = _field_rows(core_id, spec, model, verified)
    return source, _summary(core_id, cap_rows, field_rows, spec), cap_rows, field_rows


def _write_xlsx(path: Path, core_id: str, source: dict, summary, cap_rows, field_rows) -> None:
    """The half a person reads. openpyxl is imported here, not at module level: it is
    supplied per call rather than being a repo dependency, and `render_yaml` has to
    stay usable without it."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def fill(hex_colour: str) -> Any:
        return PatternFill("solid", fgColor=hex_colour)

    _GREEN, _AMBER, _RED = fill(GREEN), fill(AMBER), fill(RED)
    _GREY, _BLUE, _HEAD = fill(GREY), fill(BLUE), fill(HEAD)
    _HEAD_FONT = Font(color="FFFFFF", bold=True)
    _TITLE_FONT = Font(bold=True, size=12)

    def _head(ws: Any, row: int, headers: list[str]) -> None:
        for col, text in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=text)
            c.fill = _HEAD
            c.font = _HEAD_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)

    def _widths(ws: Any, widths: dict[int, int]) -> None:
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    wb = Workbook()

    ws = wb.active
    ws.title = "Übersicht"
    ws.cell(row=1, column=1, value=f"Prüfbogen — {core_id}").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Messlauf: {source.get('probedAt') or '—'}")
    ws.cell(row=3, column=1, value=f"Instanz: {source.get('instance') or '—'}")
    total_req = sum(s["required"] for s in summary)
    total_proven = sum(s["proven"] for s in summary)
    ws.cell(
        row=4,
        column=1,
        value=(
            f"{total_proven} von {total_req} geforderten Fähigkeiten sind live bewiesen. "
            f"„nur erreichbar“ heißt: die Route existiert und hat unsere Probe "
            f"abgelehnt — das ist kein Nachweis."
        ),
    )
    _head(
        ws,
        6,
        [
            "Entity",
            "Soll (Fähigkeiten)",
            "davon bewiesen",
            "davon unbewiesen",
            "akzeptierte Lücken",
            "nicht gebaut",
            "Feld-Anforderungen",
            "Review",
        ],
    )
    ws.freeze_panes = "A7"
    for i, s in enumerate(summary, start=7):
        for col, value in enumerate(
            [
                s["entity"],
                s["required"],
                s["proven"],
                s["unproven"],
                s["acceptedGaps"],
                s["notBuilt"],
                s["fieldRequirements"],
                s["reviewed"] or "offen",
            ],
            start=1,
        ):
            cell = ws.cell(row=i, column=col, value=value)
            if col == 3 and s["proven"]:
                cell.fill = _GREEN
            if col == 4 and s["unproven"]:
                cell.fill = _AMBER
            if col == 6 and s["notBuilt"]:
                cell.fill = _RED
            if col == 8 and not s["reviewed"]:
                cell.fill = _GREY
    ws.auto_filter.ref = f"A6:H{max(len(summary) + 6, 6)}"
    _widths(ws, {1: 22, 2: 17, 3: 15, 4: 17, 5: 18, 6: 13, 7: 18, 8: 12})

    caps = wb.create_sheet("Fähigkeiten")
    _head(caps, 1, ["Entity", "Fähigkeit", "Soll", "Ist", "Beweis", "Begründung / Hinweis"])
    caps.freeze_panes = "A2"
    for i, r in enumerate(cap_rows, start=2):
        for col, value in enumerate(
            [
                r["entity"],
                r["capability"],
                REQUIREMENT_LABEL[r["required"]],
                IMPLEMENTATION_LABEL[r["implementation"]],
                VERDICT_LABEL.get(r["verdict"], str(r["verdict"])) if r["required"] else "–",
                r["note"],
            ],
            start=1,
        ):
            cell = caps.cell(row=i, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 6))
            if col == 5:
                cell.fill = (
                    _GREEN
                    if r["proven"]
                    else _BLUE
                    if not r["required"]
                    else _RED
                    if r["verdict"] == "fail"
                    else _GREY
                    if r["verdict"] is None
                    else _AMBER
                )
            if col == 4 and r["implementation"] == "missing":
                cell.fill = _RED
    caps.auto_filter.ref = f"A1:F{max(len(cap_rows) + 1, 1)}"
    _widths(caps, {1: 20, 2: 26, 3: 20, 4: 24, 5: 30, 6: 90})

    flds = wb.create_sheet("Felder")
    _head(flds, 1, ["Entity", "Feld", "Operationen", "Status", "Fachlicher Grund", "Hinweis"])
    flds.freeze_panes = "A2"
    for i, r in enumerate(field_rows, start=2):
        hint = []
        if r["nowProven"]:
            hint.append("live bewiesen für: " + ", ".join(r["nowProven"]))
        if r["nowDeclared"]:
            hint.append("Schema deklariert: " + ", ".join(r["nowDeclared"]))
        for col, value in enumerate(
            [
                r["entity"],
                r["field"],
                ", ".join(r["ops"]),
                FIELD_STATUS_LABEL[r["status"]],
                r["reason"],
                " · ".join(hint),
            ],
            start=1,
        ):
            cell = flds.cell(row=i, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 6)))
            if col == 4:
                cell.fill = _RED if not r["fieldExists"] else _AMBER if hint else _BLUE
    flds.auto_filter.ref = f"A1:F{max(len(field_rows) + 1, 1)}"
    _widths(flds, {1: 20, 2: 34, 3: 20, 4: 26, 5: 80, 6: 34})

    wb.save(path)


def main(core_id: str = "agentos_neo_xentral") -> int:
    core_dir = _CORES / core_id
    if not (core_dir / "erp-spec.yaml").is_file():
        print(f"{core_id}: no erp-spec.yaml — nothing to review", file=sys.stderr)
        return 1

    source, summary, cap_rows, field_rows = build(core_id)
    (core_dir / "review.yaml").write_text(
        render_yaml(core_id, source, summary, cap_rows, field_rows), "utf-8"
    )
    _write_xlsx(core_dir / "review.xlsx", core_id, source, summary, cap_rows, field_rows)

    required = sum(s["required"] for s in summary)
    proven = sum(s["proven"] for s in summary)
    print(
        f"{len(summary)} Entities · {required} geforderte Fähigkeiten, davon {proven} bewiesen · "
        f"{len(field_rows)} Feld-Anforderungen → {core_dir}/review.yaml + review.xlsx"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(*sys.argv[1:2]))
