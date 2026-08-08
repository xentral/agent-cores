"""Render a core's live schema + verified.json into a readable Excel workbook.

`verified.json` records only *results*, so reading it alone cannot tell you what
was never tested — the untested fields simply are not in it. This joins the two
sides: the schema supplies the rows (what the core declares), the manifest colours
the cells (what a live run proved).

Nothing here talks to a tenant. `adapter.metadata()` already stamps everything the
workbook needs — `verified` (per-facet verdicts + `<facet>Note`), `priority` (the
blue wishes from erp-spec.yaml) and `description` — so the export is a pure
offline read of the checked-in files.

Cell vocabulary, per field × facet:

    ok       the capability ITSELF was demonstrated — a value observed, a write
             read back, an action's effect seen
    schwach  probed, but the claim could not be asserted: an HTTP 200 with nothing
             checked beyond the status, a declared field no record carried a value
             for, an action that ran without its effect being read back
    FEHLER   probed and failed — the note says why
    offen    the core declares the facet, but no run has looked at it
    –        the schema does not declare the facet for this field (not applicable)
    Wunsch   a blue wish: declared as not-possible, with a reason

Those distinctions are the point of the sheet. "offen" and "–" look the same in the
JSON (both absent) and mean opposite things; and "schwach" used to be painted "ok",
which is how 1218 read cells and 34 of 62 action cells came to claim a capability
nobody had shown. See `cores/agentos_neo_xentral/verdicts.py` for the vocabulary.

"schwach" is not automatically a to-do. Some of it is the final answer: `send`
mails a real customer, so its effect cannot be verified on a live tenant.

Run (openpyxl is deliberately not a repo dependency — supplied per call, like the
test runs supply pytest):

    PYTHONPATH=<agent-os>/backend \\
      uv run --project <agent-os>/backend --with openpyxl \\
      python scripts/export_verified_xlsx.py [core_id]
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
import types
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parent.parent
_CORES = _ROOT / "cores"

# Cores import as `xentral_entity_cores.<id>` (never top-level — a core id like
# `xentral_api` would shadow a same-named backend module). Mirrors conftest.py.
_PKG = "xentral_entity_cores"
if _PKG not in sys.modules:
    _pkg = types.ModuleType(_PKG)
    _pkg.__package__ = _PKG
    _pkg.__path__ = [str(_CORES)]  # type: ignore[attr-defined]
    sys.modules[_PKG] = _pkg

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from xentral_entity_cores.agentos_neo_xentral.verdicts import (  # noqa: E402
    VERDICTS,
    is_proven,
)

# The facet vocabulary the runner records, in the order the backend surfaces it
# (entity_registry/verification.py::CAPABILITY_FACETS).
FACETS = ("read", "create", "update", "filter", "sort", "search")

# facet → the schema flag that says whether it applies to a field at all.
# `read` has no flag: everything declared is readable, including readOnly fields.
_FLAG = {
    "create": "creatable",
    "update": "updatable",
    "filter": "filterable",
    "sort": "sortable",
    "search": "searchable",
}

OK, FAIL, OPEN, NA, WISH = "ok", "FEHLER", "offen", "–", "Wunsch"
# Probed, but the run could not assert the claim itself: an HTTP 200 with nothing
# checked beyond the status, a declared field no record carried a value for, an
# action that ran without its effect being read back. Neither green nor a defect —
# for some of them (`send` mails a real customer) it is the final answer, not a
# to-do. Distinct from `offen`, which means no run has looked at all.
WEAK = "schwach"

_FILL = {
    OK: PatternFill("solid", fgColor="C6E0B4"),
    FAIL: PatternFill("solid", fgColor="F4B183"),
    WEAK: PatternFill("solid", fgColor="FFE699"),
    OPEN: PatternFill("solid", fgColor="EDEDED"),
    NA: PatternFill("solid", fgColor="FFFFFF"),
    WISH: PatternFill("solid", fgColor="BDD7EE"),
}
_HEAD = PatternFill("solid", fgColor="44546A")
_HEAD_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=12)


def _cell(spec: dict[str, Any], facet: str) -> tuple[str, str | None]:
    """One field × facet → (value, note)."""
    wish = (spec.get("priority") or {}).get(facet)
    if wish:
        return WISH, wish
    verified = spec.get("verified") or {}
    note = verified.get(f"{facet}Note")
    verdict = verified.get(facet)
    if is_proven(verdict):
        return OK, note
    if verdict == "fail":
        return FAIL, note or "probe failed"
    if verdict in VERDICTS:
        # Probed, but the claim was not asserted. The note carries which of the
        # weak verdicts it was and why, so it is never dropped here.
        return WEAK, note or verdict
    flag = _FLAG.get(facet)
    if flag and not spec.get(flag):
        # Not declared for this field — asking whether it was tested is meaningless.
        # A leftover note here is worth keeping though: the runner only writes one
        # for a facet it considered applicable, so it means the schema has since
        # dropped the flag and the manifest is stale on that cell.
        return NA, (f"veraltet? {note}" if note else None)
    return OPEN, note


def _status(entry: dict[str, Any]) -> tuple[str, str | None]:
    """One action or process-step command → (cell value, note).

    The action twin of `_cell`. Module level rather than a closure so the two can be
    tested against the same vocabulary — they drifted once already, and the drift is
    invisible in the sheet.
    """
    if entry.get("wish"):
        return WISH, entry["wish"]
    verified = entry.get("verified") or {}
    v = verified.get("status")
    # The note is what separates "we ran it and saw the effect" from "the route
    # refused our empty probe". Dropping it made 34 of 62 action cells read as proof
    # of a capability that was never exercised.
    note = verified.get("note")
    if is_proven(v):
        return OK, note
    if v == "fail":
        return FAIL, note
    if v in VERDICTS:
        return WEAK, note or v
    return OPEN, note


def _walk(props: Any, prefix: str = "") -> list[tuple[str, dict[str, Any]]]:
    """Every declared path, containers included, so `addresses` and
    `addresses.city` both get a row. Recurses embedded `properties` AND collection
    `node.properties` — the same shape verify.py::_field_paths walks."""
    out: list[tuple[str, dict[str, Any]]] = []
    if not isinstance(props, dict):
        return out
    for name, spec in sorted(props.items()):
        if not isinstance(spec, dict):
            continue
        path = f"{prefix}.{name}" if prefix else name
        out.append((path, spec))
        sub = spec.get("properties")
        if not isinstance(sub, dict):
            node = spec.get("node")
            sub = node.get("properties") if isinstance(node, dict) else None
        out += _walk(sub, path)
    return out


_BAD_TITLE = re.compile(r"[\[\]:*?/\\]")


def _when(stamp: Any) -> str:
    """An ISO stamp as `YYYY-MM-DD HH:MM`, empty when never measured.

    A manifest is assembled from many scoped runs, so its entities genuinely have
    different ages — the seven sales documents were re-probed weeks after the rest.
    One date for the whole file would hide that.
    """
    if not isinstance(stamp, str) or not stamp:
        return ""
    try:
        return dt.datetime.fromisoformat(stamp).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return stamp[:16]


def _action_when(manifest: dict[str, Any], key: str, verdicts: int) -> str:
    """Empty means "never probed", `?` means "probed before the run stamped itself".
    Blank for both would read as never, which is a different claim."""
    stamp = ((manifest.get("entities") or {}).get(key) or {}).get("actionsProbedAt")
    when = _when(stamp)
    if when:
        return when
    return "?" if verdicts else ""


def _sheet_title(key: str, used: set[str]) -> str:
    """Excel: max 31 chars, no []:*?/\\ , unique. The full key stays on the
    overview sheet, so a shortened tab never loses information."""
    title = _BAD_TITLE.sub("-", key)[:31]
    if title not in used:
        used.add(title)
        return title
    for n in range(2, 100):
        cand = f"{title[: 31 - len(str(n)) - 1]}~{n}"
        if cand not in used:
            used.add(cand)
            return cand
    raise RuntimeError(f"no unique sheet title for {key}")


def _write_header(ws: Any, row: int, headers: list[str]) -> None:
    for col, head in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=head)
        c.fill = _HEAD
        c.font = _HEAD_FONT
        c.alignment = Alignment(vertical="center")


def _autosize(ws: Any, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _entity_sheet(
    wb: Workbook, key: str, meta: dict[str, Any], title: str, probed_at: str = ""
) -> dict[str, Any]:
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=f"{key} — {meta.get('label') or ''}").font = _TITLE_FONT
    ws.cell(row=2, column=1, value="Operations: " + ", ".join(meta.get("operations") or []))
    if probed_at:
        ws.cell(row=2, column=4, value=f"gemessen: {probed_at}")

    headers = ["Pfad", "Typ", "Label", *FACETS, "nur Einzelabruf", "Beschreibung", "Notizen"]
    _write_header(ws, 4, headers)
    ws.freeze_panes = "A5"

    props = (meta.get("rootNode") or {}).get("properties") or {}
    tally = {v: 0 for v in (OK, WEAK, FAIL, OPEN, NA, WISH)}
    row = 5
    for path, spec in _walk(props):
        notes: list[str] = []
        ws.cell(row=row, column=1, value=path)
        ws.cell(row=row, column=2, value=spec.get("type"))
        ws.cell(row=row, column=3, value=spec.get("label"))
        for i, facet in enumerate(FACETS):
            value, note = _cell(spec, facet)
            tally[value] += 1
            c = ws.cell(row=row, column=4 + i, value=value)
            c.fill = _FILL[value]
            c.alignment = Alignment(horizontal="center")
            if note:
                notes.append(f"{facet}: {note}")
        # Filled from a sub-resource the core reads only for a single record — a
        # list leaves it null. Its own column, because "the value is null here for
        # a structural reason" is not a test result and does not belong in a facet.
        ws.cell(row=row, column=10, value="ja" if spec.get("detailOnly") else "")
        ws.cell(row=row, column=11, value=spec.get("description"))
        ws.cell(row=row, column=12, value=" | ".join(notes))
        row += 1

    field_rows = row - 5
    ws.auto_filter.ref = f"A4:L{max(row - 1, 4)}"

    def _block(start: int, heading: str, cols: list[str], rows: list[list[Any]]) -> int:
        ws.cell(row=start, column=1, value=heading).font = _TITLE_FONT
        _write_header(ws, start + 1, cols)
        r = start + 2
        for values in rows:
            for ci, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                if ci == len(cols) - 1 and v in _FILL:
                    c.fill = _FILL[v]
                    c.alignment = Alignment(horizontal="center")
            r += 1
        if not rows:
            ws.cell(row=r, column=1, value="(keine)")
            r += 1
        return r + 1

    act_rows = []
    for a in meta.get("actions") or []:
        status, note = _status(a)
        act_rows.append(
            [
                a.get("key"),
                a.get("label"),
                "ja" if a.get("destructive") else "",
                a.get("description") or "",
                status,
                note or "",
            ]
        )
    nxt = _block(
        row + 1,
        "Actions",
        ["Key", "Label", "destruktiv", "Beschreibung", "Status", "Notiz"],
        act_rows,
    )

    step_rows = []
    for grp in meta.get("processSteps") or []:
        for cmd in (grp or {}).get("commands") or []:
            status, note = _status(cmd)
            step_rows.append(
                [
                    cmd.get("key"),
                    f"{grp.get('label') or grp.get('group')} › {cmd.get('label')}",
                    "ja" if cmd.get("destructive") else "",
                    cmd.get("description") or "",
                    status,
                    note or "",
                ]
            )
    last = _block(
        nxt,
        "Process-Steps",
        ["Key", "Label", "destruktiv", "Beschreibung", "Status", "Notiz"],
        step_rows,
    )

    # Wishes for a field the schema does not have. They have no row to colour blue
    # above — "items.totals.tax" is missing precisely BECAUSE upstream states no
    # per-line tax amount — so without their own block the widest gaps in the
    # backlog would be the only ones invisible in this workbook.
    missing = meta.get("missingFieldWishes") or []
    if missing:
        _block(
            last,
            "Wünsche ohne Feld (Lücke im Modell)",
            ["Feld", "Facetten", "Status", "Begründung"],
            [
                [m.get("field"), ", ".join(m.get("ops") or []), WISH, m.get("reason") or ""]
                for m in missing
            ],
        )

    _autosize(
        ws,
        {1: 34, 2: 12, 3: 22, 4: 8, 5: 8, 6: 8, 7: 8, 8: 8, 9: 8, 10: 14, 11: 60, 12: 70},
    )

    def _count(rows: list[list[Any]], value: str) -> int:
        return sum(1 for r in rows if r[4] == value)

    return {
        "fields": field_rows,
        "tally": tally,
        "actions": (len(act_rows), _count(act_rows, OK), _count(act_rows, FAIL)),
        "steps": (len(step_rows), _count(step_rows, OK), _count(step_rows, FAIL)),
    }


def main(core_id: str = "agentos_neo_xentral") -> int:
    core_pkg = __import__(f"{_PKG}.{core_id}", fromlist=["CORE"])
    core = core_pkg.CORE

    manifest_path = _CORES / core_id / "verified.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    # The run stamps itself since PR #74. Older manifests carry generatedAt: null —
    # then the file's mtime is all there is, and it is labelled as the guess it is
    # (after a fresh clone the mtime is the checkout time).
    generated = manifest.get("generatedAt")
    if generated:
        stamped = f"Lauf beendet: {_when(generated)}"
    else:
        mtime = dt.datetime.fromtimestamp(manifest_path.stat().st_mtime)
        stamped = f"kein Laufstempel — Dateizeit: {mtime.strftime('%Y-%m-%d %H:%M')}"
    probed = {k: (v or {}).get("probedAt") for k, v in (manifest.get("entities") or {}).items()}

    wb = Workbook()
    overview = wb.active
    overview.title = "Übersicht"

    used: set[str] = set()
    rows: list[tuple[str, str, dict[str, Any]]] = []
    for adapter in core.adapters:
        key = adapter.manifest.key
        meta = adapter.metadata()
        title = _sheet_title(key, used)
        rows.append((key, title, _entity_sheet(wb, key, meta, title, _when(probed.get(key)))))

    overview.cell(row=1, column=1, value=f"Kern: {core_id}").font = _TITLE_FONT
    overview.cell(row=2, column=1, value=f"Instanz: {manifest.get('instance') or '?'}")
    overview.cell(row=3, column=1, value=stamped)
    overview.cell(
        row=4,
        column=1,
        value=(
            "ok = die Fähigkeit selbst wurde nachgewiesen · "
            "schwach = geprüft, aber der Nachweis blieb aus (HTTP 200 ohne Wirkungs"
            "prüfung, Feld nie befüllt gesehen, Action ohne Effektkontrolle) — bei "
            "manchen, z. B. send, ist das die endgültige Antwort · "
            "FEHLER = geprüft und fehlgeschlagen · "
            "offen = deklariert, aber ungeprüft · – = laut Schema nicht anwendbar · "
            "Wunsch = bewusst nicht möglich (erp-spec.yaml)"
        ),
    )

    headers = [
        "Entity",
        "Tab",
        "Felder",
        "ok",
        "schwach",
        "FEHLER",
        "offen",
        "–",
        "Wunsch",
        "Actions (ok/Fehler)",
        "Steps (ok/Fehler)",
        "gemessen",
        "Actions gemessen",
    ]
    _write_header(overview, 6, headers)
    overview.freeze_panes = "A7"
    r = 7
    for key, title, stats in sorted(rows):
        t = stats["tally"]
        a_all, a_ok, a_fail = stats["actions"]
        s_all, s_ok, s_fail = stats["steps"]
        for ci, v in enumerate(
            [
                key,
                title,
                stats["fields"],
                t[OK],
                t[WEAK],
                t[FAIL],
                t[OPEN],
                t[NA],
                t[WISH],
                f"{a_all} ({a_ok}/{a_fail})",
                f"{s_all} ({s_ok}/{s_fail})",
                _when(probed.get(key)),
                _action_when(manifest, key, a_ok + a_fail),
            ],
            start=1,
        ):
            overview.cell(row=r, column=ci, value=v)
        r += 1
    overview.auto_filter.ref = f"A6:K{max(r - 1, 6)}"
    _autosize(
        overview,
        {1: 28, 2: 28, 3: 9, 4: 8, 5: 10, 6: 9, 7: 8, 8: 6, 9: 9, 10: 20, 11: 20, 12: 17, 13: 17},
    )

    out = _CORES / core_id / "verified.xlsx"
    wb.save(out)
    total = sum(s["fields"] for _, _, s in rows)
    fails = sum(s["tally"][FAIL] for _, _, s in rows)
    print(f"{len(rows)} Entities, {total} Feldzeilen, {fails} FEHLER-Zellen → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(*sys.argv[1:2]))
